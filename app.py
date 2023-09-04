from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
import openpyxl

# Reduz as funções


def find(xpath):
    element = driver.find_element(By.XPATH, f"{xpath}")
    return element


def click(xpath):
    element = find(xpath)
    element.click()
    return element


def click_nextChapter(xpath, iterative, totalChapters):
    if iterative < totalChapters:
        click(xpath)
    else:
        pass


# Processamento de Dados
pontuacoes = [
    ".",
    ",",
    ":",
    ";",
    "!",
    "?",
    "[",
    "]",
    "(",
    ")",
    '"',
    "'",
    "“",
    "”",
    "’",
    "`",
    "—",
    "–",
    "-",
    "/",
    "*",
    "\n",
    "  ",
]


def removerPontuacao(content):
    for pontuacao in pontuacoes:
        content = content.replace(pontuacao, "")
    return content


def removerEspaco(content):
    content = content.replace("\n", "")
    content = content.replace(" ", "")
    return content


def extrairDados(text):
    dadosProcessados = []
    dadosProcessados.append(len(text.split()))
    dadosProcessados.append(len(text.replace("\n", "")))
    dadosProcessados.append(len(removerEspaco(text)))
    dadosProcessados.append(len(removerPontuacao(text)))
    dadosProcessados.append(len(removerPontuacao(removerEspaco(text))))
    return dadosProcessados


def calcularTotal(dadosExtraidos):
    lista_somaTotal = []
    for x in range(0, len(dadosExtraidos[0])):
        somaTotal = 0
        for y in range(0, len(dadosExtraidos)):
            somaTotal += dadosExtraidos[y][x]
        lista_somaTotal.append(somaTotal)
    return lista_somaTotal


def calcularMedia(dadosExtraidos):
    lista_somaTotal = calcularTotal(dadosExtraidos)
    lista_media = []
    for i in range(0, len(dadosExtraidos[0])):
        lista_media.append(lista_somaTotal[i] / len(dadosExtraidos))
    return lista_media


# Abre o navegador no site
driver = webdriver.Edge()
novels = {'ISSTH': "https://readnovelfull.me/i-shall-seal-the-heavens/",
          'RI': "https://readnovelfull.me/renegade-immortal/"}
novel_name = 'RI'
novel_link = novels[novel_name]

driver.get(novel_link)
driver.minimize_window()

# Entra no primeiro capítulo
xpaths_iniciais = [
    "//a[@id='tab-chapters-title']",
    "//*[@id='panel-book-0']/div/div/div[1]/ul/li[1]/a",
    "//*[@id='chr-nav-top']/div/button",
]
for xpath in xpaths_iniciais:
    click(xpath)
    sleep(1)

# Define a quantidade de capítulos, inicia contagem de caracteres e guarda em uma lista
totalChapters = len(driver.find_elements(
    By.XPATH, '//*[@id="chr-nav-top"]/div/select/option'))
dadosExtraidos = []

for i in range(1, totalChapters+1):
    dadosExtraidos.append(extrairDados(find("//*[@id='chr-content']").text))
    click_nextChapter('//*[@id="next_chap"]', i, totalChapters)

# Excel
workbook = openpyxl.load_workbook('dados_novel.xlsx')

# Código para criar uma página do zero e inserir as informações
paginaNovel = workbook[novel_name]
paginaNovel['A1'].value = "Capítulos"
paginaNovel['A2'].value = "Total"
paginaNovel['A3'].value = "Média"
paginaNovel['B1'].value = "Palavras"
paginaNovel['C1'].value = "Caracteres"
paginaNovel['D1'].value = "Sem espaços"
paginaNovel['E1'].value = "Sem pontuação"
paginaNovel['F1'].value = "Sem ambos"

for index, linha in enumerate(paginaNovel.iter_rows(min_row=2, min_col=2, max_row=3, max_col=6)):
    for celula in linha:
        if index == 0:
            celula.value = calcularTotal(dadosExtraidos)[linha.index(celula)]
        else:
            celula.value = calcularMedia(dadosExtraidos)[linha.index(celula)]

for index, linha in enumerate(paginaNovel.iter_rows(min_row=4, min_col=1, max_row=len(dadosExtraidos) + 3, max_col=1)):
    linha[0].value = index + 1

for index, linha in enumerate(paginaNovel.iter_rows(min_row=4, min_col=2, max_row=len(dadosExtraidos) + 3, max_col=6)):
    for celula in linha:
        celula.value = dadosExtraidos[index][linha.index(celula)]

workbook.save('dados_novel.xlsx')
driver.quit()
